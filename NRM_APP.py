import customtkinter as ctk
import pandas as pd
import json
import os
from tkinter import filedialog, messagebox, simpledialog
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from datetime import datetime
import win32com.client as win32
from openpyxl.styles import NamedStyle, Font, PatternFill
import comtypes.client
from CTkMenuBar import *
import time
import pyautogui as pg
import subprocess


class ScrollableFrame(ctk.CTkFrame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        canvas = ctk.CTkCanvas(self)
        scrollbar = ctk.CTkScrollbar(self, orientation="vertical", command=canvas.yview)
        self.scrollable_frame = ctk.CTkFrame(canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
       
class MenuFrame(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master
        # Set the width and height for the frame
        self.height = 30        
        # Configure the frame's width and height
        self.configure(height=self.height)
        #self.pack(fill="x")
                
        self.menu = CTkMenuBar(master)
        self.button_1 = self.menu.add_cascade("Tools")
        self.button_2 = self.menu.add_cascade("Formulate")
        self.button_3 = self.menu.add_cascade("Settings")
        self.button_4 = self.menu.add_cascade("About")
      
        self.dropdown1 = CustomDropdownMenu(widget=self.button_1)
        self.dropdown1.add_option(option="Add Isntruction", command=lambda: self.add_instruction())
        self.dropdown1.add_option(option="Repeat command", command=self.repeat_command)

        self.dropdown1.add_separator()   

        self.dropdown2 = CustomDropdownMenu(widget=self.button_2)
        self.dropdown2.add_option(option="ADD DB", command=lambda: self.add_db())
        self.dropdown2.add_option(option="Bar tender", command=lambda: self.formulate_btd())
        self.dropdown2.add_option(option="Corel Draw", command=self.formulate_crl)
        self.dropdown2.add_option(option="print&import", command=self.print_import)


        self.dropdown2.add_separator() 
        
    def add_instruction(self):
        new_window = ctk.CTkToplevel(self)
        new_window.title('Add Instruction')
        new_window.geometry('200x400')
        
        
        self.output_text = ctk.CTkTextbox(new_window, width=190, height=300)
        self.output_text.pack(pady=10)
        
        self.submit_inst= ctk.CTkButton(new_window, text = 'Submit', command = self.submit_instruction)
        self.submit_inst.pack(pady = 10)
        
        
    def submit_instruction(self):
        # Initialize an empty dictionary for instructions
        instructions = {}
        with open(r'\\deepa\d\MAIL-2024\ExcelFilterApp\inst.json') as file:
            instructions = json.load(file)
            print(instructions)
        
        
        # Retrieve values from the variables
        brand = self.master.brand_var.get()
        party = self.master.party_var.get()
        artwork = self.master.artwork_var.get()
        instruction_text = self.output_text.get("1.0", "end-1c")  # Replace with actual instruction text
        
        # Create the nested structure
        if brand not in instructions:
            instructions[brand] = {}  # Create a dictionary for the brand if it doesn't exist
        if party not in instructions[brand]:
            instructions[brand][party] = {}  # Create a dictionary for the party under the brand
        if artwork not in instructions[brand][party]:
            instructions[brand][party][artwork] = []  # Create a list for instructions under the artwork
        
        # Append the instruction text to the list
        instructions[brand][party][artwork].append(instruction_text)
        
        # Print the resulting structure (for debugging)
        print(instructions)
        
        with open(r'\\deepa\d\MAIL-2024\ExcelFilterApp\inst.json', 'w', encoding='utf-8') as json_file:
            json.dump(instructions, json_file, ensure_ascii=False, indent=4)
    
    def repeat_command(self):
        method_handler = MethodsHandler(self.master)
        method_handler.repeat_command()       
    
    def print_import(self):
        method_handler = MethodsHandler(self.master)
        method_handler.print_import()
    
    def add_db(self):
        method_handler = MethodsHandler(self.master)
        method_handler.add_db()
        
    def formulate_btd(self):
        method_handler = MethodsHandler(self.master)
        method_handler.formulate_bartender()
        
    def formulate_crl(self):
        method_handler = MethodsHandler(self.master)
        method_handler.formulate_corel()
        

class ExcelFilterApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Excel Filter App")
        self.geometry("400x600")
        # MENUBAR FOR ADDING MENUS
        
        menu = MenuFrame(self)
        
        #self.menu_bar = ctk.CTkFrame(self, height=30)
        #self.menu_bar.pack(fill="x")
        
        # Opeining window for adding the details in excel
                
        self.scrollable_frame = ScrollableFrame(self)
        self.scrollable_frame.pack(fill="both", expand=True)

        # Brand selection
        self.brand_label = ctk.CTkLabel(self.scrollable_frame.scrollable_frame, text="Select Brand:")
        self.brand_label.pack(pady=1)
        self.brand_var = ctk.StringVar(value="Select Brand")
        self.brand_optionmenu = ctk.CTkOptionMenu(self.scrollable_frame.scrollable_frame, variable=self.brand_var, values=[], command=self.update_parties)
        self.brand_optionmenu.pack(pady=1)

        self.add_brand_button = ctk.CTkButton(self.scrollable_frame.scrollable_frame, text="Add Brand", command=self.add_brand)
        self.add_brand_button.pack(pady=1)

        # Party selection
        self.party_label = ctk.CTkLabel(self.scrollable_frame.scrollable_frame, text="Select Party:")
        self.party_label.pack(pady=1)
        self.party_var = ctk.StringVar(value="Select Party")
        self.party_optionmenu = ctk.CTkOptionMenu(self.scrollable_frame.scrollable_frame, variable=self.party_var, values=[], command=self.update_artworks)
        self.party_optionmenu.pack(pady=1)

        self.add_party_button = ctk.CTkButton(self.scrollable_frame.scrollable_frame, text="Add Party", command=self.add_party)
        self.add_party_button.pack(pady=1)

        # Artwork type selection
        self.artwork_label = ctk.CTkLabel(self.scrollable_frame.scrollable_frame, text="Select Artwork Type:")
        self.artwork_label.pack(pady=1)
        self.artwork_var = ctk.StringVar(value="Select Artwork Type")
        self.artwork_optionmenu = ctk.CTkOptionMenu(self.scrollable_frame.scrollable_frame, variable=self.artwork_var, values=[], command=self.load_selected_settings)
        self.artwork_optionmenu.pack(pady=1)

        self.add_artwork_button = ctk.CTkButton(self.scrollable_frame.scrollable_frame, text="Add Artwork Type", command=self.add_artwork)
        self.add_artwork_button.pack(pady=1)

        # Default directories and file paths
        self.input_dir_button = ctk.CTkButton(self.scrollable_frame.scrollable_frame, text="Select Default Input Directory", command=self.set_default_input_dir)
        self.input_dir_button.pack(pady=10)
        self.input_dir_label = ctk.CTkLabel(self.scrollable_frame.scrollable_frame, text="")
        self.input_dir_label.pack(pady=1)

        self.output_dir_button = ctk.CTkButton(self.scrollable_frame.scrollable_frame, text="Select Default Output Directory", command=self.set_default_output_dir)
        self.output_dir_button.pack(pady=10)
        self.output_dir_label = ctk.CTkLabel(self.scrollable_frame.scrollable_frame, text="")
        self.output_dir_label.pack(pady=1)

        self.corel_file_button = ctk.CTkButton(self.scrollable_frame.scrollable_frame, text="Select Sample CorelDRAW File", command=self.set_corel_file)
        self.corel_file_button.pack(pady=10)
        self.corel_file_label = ctk.CTkLabel(self.scrollable_frame.scrollable_frame, text="")
        self.corel_file_label.pack(pady=1)

        self.load_sample_button = ctk.CTkButton(self.scrollable_frame.scrollable_frame, text="Load Sample File", command=self.load_sample_file)
        self.load_sample_button.pack(pady=10)
        self.sample_file_label = ctk.CTkLabel(self.scrollable_frame.scrollable_frame, text="")
        self.sample_file_label.pack(pady=1)

        # Load Excel File button
        self.load_button = ctk.CTkButton(self.scrollable_frame.scrollable_frame, text="Load Excel File to Filter", command=self.load_excel_file)
        self.load_button.pack(pady=10)
        self.file_label = ctk.CTkLabel(self.scrollable_frame.scrollable_frame, text="")
        self.file_label.pack(pady=1)

        self.filter_button = ctk.CTkButton(self.scrollable_frame.scrollable_frame, text="Filter and Save Excel File", command=self.filter_and_save)
        self.filter_button.pack(pady=20)

        self.sample_data = None
        self.excel_data = None
        self.required_columns = []
        self.column_mapping = {}
        self.settings_dir = r'\\DEEPA\D\MAIL-2024\ExcelFilterApp'
        self.json_file = os.path.join(self.settings_dir, 'settings.json')
        self.inst_dir= r'\\deepa\d\MAIL-2024\ExcelFilterApp'
        self.inst_file = os.path.join(self.inst_dir, 'inst.json')

        self.load_settings()
        
    
    def load_instuctions(self):
        with open(self.inst_file) as file:
            instructions = json.load(file)
            
        brand = self.brand_var.get()
        party = self.party_var.get()
        artwork = self.artwork_var.get()
        brand_data = instructions.get(brand, {})
        party_data = brand_data.get(party, {})
        artwork_data = party_data.get(artwork, {})
        length = len(artwork_data)
        if length:
            message = ""
            i = 0
            for item in artwork_data:
                i = i+1
                message = message + f"{i}: {item}\n"
            messagebox.showinfo("Special instructions", message)
        
    def read_excel(self, cdr_file_path):
        brand = self.brand_var.get()
        party = self.party_var.get()
        artwork = self.artwork_var.get()
        path = r"\\Deepa\d\MAIL-2024\ExcelFilterApp\DAILY_REPOS.xlsx"
        
        # Read the existing Excel file into a DataFrame
        df = pd.read_excel(path)
        
        # Create a new DataFrame with the data to append
        new_data = pd.DataFrame({
            'PARTY NAME': [party],
            'BRAND': [brand],
            'ARTWORK': [artwork],
            'CDR File':[cdr_file_path]
        })
        
        # Concatenate new_data with the existing DataFrame
        df = pd.concat([df, new_data], ignore_index=True)
                
        # Display confirmation dialog before saving
        confirm = messagebox.askyesno("Confirm Save", "Do you want to save the changes?")
        
        if confirm:
            # Save the DataFrame back to the same Excel file
            df.to_excel(path, index=False)
            messagebox.showinfo("Save Successful", f"DataFrame saved to {path}")
        else:
            messagebox.showinfo("Save Canceled", "Changes were not saved.")

    def load_settings(self):
        try:
            if not os.path.exists(self.settings_dir):
                os.makedirs(self.settings_dir)
            with open(self.json_file, 'r') as f:
                settings = json.load(f)
                self.settings = settings
                self.brand_optionmenu.configure(values=list(settings.keys()))
        except (FileNotFoundError, KeyError, PermissionError) as e:
            self.settings = {}
            self.initialize_default_settings()
            if isinstance(e, PermissionError):
                messagebox.showwarning("Warning", f"Failed to load settings file due to permission error: {e}")

    def initialize_default_settings(self):
        self.brand_var.set('easybuy')
        self.party_var.set('jc hometex')
        self.artwork_var.set('washcare')
        self.sample_file_label.configure(text='No sample file loaded')
        self.input_dir_label.configure(text=r'\\deepa\d\MAIL-2024')
        self.output_dir_label.configure(text=r'\\deepa\d\nrm-2024')
        self.corel_file_label.configure(text='No CorelDRAW file set')
        self.column_mapping = {
            'EAN': ['EAN code', 'EAN number', 'Barcode', 'item code'],
            'Style Number': ['Style No', 'Style ID'],
            'Vendor Code': ['Vendor ID'],
            'Vendor Name': ['Vendor'],
            'MRP': ['Price'],
            'Content': ['Material'],
            'Wash Inst': ['Washing Instructions', 'Care Instructions']
        }

    def save_settings(self):
        brand = self.brand_var.get()
        party = self.party_var.get()
        artwork = self.artwork_var.get()

        if brand not in self.settings:
            self.settings[brand] = {}

        if party not in self.settings[brand]:
            self.settings[brand][party] = {}

        if artwork not in self.settings[brand][party]:
            self.settings[brand][party][artwork] = {}

        self.settings[brand][party][artwork] = {
            'sample_file': self.sample_file_label.cget("text"),
            'input_dir': self.input_dir_label.cget("text"),
            'output_dir': self.output_dir_label.cget("text"),
            'corel_file': self.corel_file_label.cget("text"),
            'column_mapping': self.column_mapping
        }

        try:
            with open(self.json_file, 'w') as f:
                json.dump(self.settings, f)
        except PermissionError as e:
            messagebox.showwarning("Warning", f"Failed to save settings file due to permission error: {e}")

    def update_parties(self, selected_brand):
        parties = list(self.settings.get(selected_brand, {}).keys())
        #print(f'found these parties {parties}')
        if parties:
            self.party_optionmenu.configure(values=parties)
            self.party_var.set(parties[0])
            self.load_selected_settings()
    def update_artworks(self, selected_party):
        brand = self.brand_var.get()
        
        artworks = list(self.settings.get(brand, {}).get(selected_party, {}).keys())
        #print(f'found these parties {artworks}')
        if artworks:
            self.artwork_optionmenu.configure(values=artworks)
            self.artwork_var.set(artworks[0])
            self.load_selected_settings()        
        
    def load_selected_settings(self, *_):
        brand = self.brand_var.get()
        #print(f'selected brand : {brand}')
        party = self.party_var.get()
        #print(f'Updated party : {party}')
        artwork = self.artwork_var.get()
        #print(f'updated artwork {artwork}')

        settings = self.settings.get(brand, {}).get(party, {}).get(artwork, {})
        #print(settings)
        sample_file = settings.get('sample_file', '')
        if sample_file:
            self.sample_file_label.configure(text=sample_file)
            self.sample_data = pd.read_excel(sample_file)
            self.required_columns = self.sample_data.columns.tolist()
        self.input_dir_label.configure(text=settings.get('input_dir', 'No input directory set'))
        self.output_dir_label.configure(text=settings.get('output_dir', 'No output directory set'))
        self.corel_file_label.configure(text=settings.get('corel_file', 'No CorelDRAW file set'))
        self.column_mapping = settings.get('column_mapping', self.column_mapping)

    def set_default_input_dir(self):
        input_dir = filedialog.askdirectory(initialdir = self.input_dir_label.cget("text"),)
        if input_dir:
            self.input_dir_label.configure(text=input_dir)
            self.save_settings()

    def set_default_output_dir(self):
        output_dir = filedialog.askdirectory( initialdir = self.output_dir_label.cget("text"),)
        if output_dir:
            self.output_dir_label.configure(text=output_dir)
            self.save_settings()

    def set_corel_file(self):
        corel_file = filedialog.askopenfilename(filetypes=[("CorelDRAW files", "*.cdr")])
        if corel_file:
            self.corel_file_label.configure(text=corel_file)
            self.save_settings()

    def load_sample_file(self):
        sample_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if sample_file_path:
            self.sample_file_label.configure(text=sample_file_path)
            self.sample_data = pd.read_excel(sample_file_path)
            self.required_columns = self.sample_data.columns.tolist()
            self.save_settings()

    def load_excel_file(self):
        file_path = filedialog.askopenfilename(initialdir = self.input_dir_label.cget("text"), filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            self.file_label.configure(text=file_path)
            self.excel_data = pd.read_excel(file_path)

    def add_brand(self):
        new_brand = simpledialog.askstring("Add Brand", "Enter new brand name:")
        if new_brand:
            if new_brand not in self.settings:
                self.settings[new_brand] = {}
                self.brand_optionmenu.configure(values=list(self.settings.keys()))
                self.save_settings()
            self.brand_var.set(new_brand)

    def add_party(self):
        brand = self.brand_var.get()
        if brand == "Select Brand":
            messagebox.showerror("Error", "Please select a brand first.")
            return
        new_party = simpledialog.askstring("Add Party", "Enter new party name:")
        if new_party:
            if new_party not in self.settings[brand]:
                self.settings[brand][new_party] = {}
                self.party_optionmenu.configure(values=list(self.settings[brand].keys()))
                self.save_settings()
            self.party_var.set(new_party)

    def add_artwork(self):
        brand = self.brand_var.get()
        party = self.party_var.get()
        if brand == "Select Brand" or party == "Select Party":
            messagebox.showerror("Error", "Please select a brand and party first.")
            return
        new_artwork = simpledialog.askstring("Add Artwork Type", "Enter new artwork type:")
        if new_artwork:
            if new_artwork not in self.settings[brand][party]:
                self.settings[brand][party][new_artwork] = {}
                self.artwork_optionmenu.configure(values=list(self.settings[brand][party].keys()))
                self.save_settings()
            self.artwork_var.set(new_artwork)
            
    def change_extension(self, file_path):
        base_name, _ = os.path.splitext(file_path)
        new_path = base_name + ".cdr"
        
        return new_path

    def create_blank_cdr_file(self, output_path):
        
        try:
            # Initialize CorelDRAW application
            coreldraw = comtypes.client.CreateObject("CorelDRAW.Application")

            # Create a new document
            doc = coreldraw.CreateDocument()

            # Save the document as a .cdr file
            if not os.path.isabs(output_path):
                output_path = os.path.abspath(output_path)  # Ensure the path is absolute

            # Save the document
            doc.SaveAs(output_path)
            print(f"Document saved as: {output_path}")

            # Debug: Print list of open documents
            open_docs = coreldraw.Documents
            print(f"Number of open documents before closing: {len(open_docs)}")

            # Close all open documents
            for document in open_docs:
                try:
                    print(f"Closing document: {document.Name}")
                    document.Close()
                except Exception as close_error:
                    print(f"Error closing document {document.Name}: {close_error}")
                    
            import win32com.client
            # Replace forward slashes with backward slashes
            output_path = output_path.replace('/', '\\')
            
            # Connect to CorelDRAW application
            corel_app = win32com.client.Dispatch("CorelDRAW.Application")
                    
            corel_doc = corel_app.OpenDocument(output_path)

            # Debug: Print list of open documents after closing
            open_docs_after = coreldraw.Documents
            print(f"Number of open documents after closing: {len(open_docs_after)}")

            
        except Exception as e:
            print(f"Error creating, closing, or opening the file: {e}")
    
    def filter_and_save(self):
        self.load_instuctions()
        if self.excel_data is None or self.sample_data is None:
            messagebox.showerror("Error", "Please load both sample and Excel files first.")
            return

        # List of potential column names
        potential_columns = ['STYLE DETAILS','style_code', 'Style Number', 'generic_number', 'generic_code','STYLE','SHORT_DESC', 'Style Code', 'Generic/Single article code','SHORT_ DESC', 'STYLE CODE','Stylecode', 'Vendor Part number', 'PO NO', 'PO No.']

        brand = self.brand_var.get()
        party = self.party_var.get()
        artwork = self.artwork_var.get()

        column_mapping = self.settings.get(brand, {}).get(party, {}).get(artwork, {}).get('column_mapping', {})
        missing_columns = []
        for required_col in self.required_columns:
            if required_col not in self.excel_data.columns:
                mapped = False
                for alt_name in column_mapping.get(required_col, []):
                    if alt_name in self.excel_data.columns:
                        self.excel_data.rename(columns={alt_name: required_col}, inplace=True)
                        mapped = True
                        break
                if not mapped:
                    missing_columns.append(required_col)

        for col in missing_columns:
            options = [col for col in self.excel_data.columns if col not in self.required_columns]
            selected_col = simpledialog.askstring("Column Mapping", f"Select column to map with {col}:\nOptions: {options}")
            if selected_col and selected_col in self.excel_data.columns:
                self.excel_data.rename(columns={selected_col: col}, inplace=True)
                column_mapping[col] = [selected_col]

        extra_columns = [col for col in self.excel_data.columns if col not in self.required_columns]
        self.excel_data.drop(columns=extra_columns, inplace=True)

        self.settings[brand][party][artwork]['column_mapping'] = column_mapping
        self.save_settings()
        
        formulae_handler = FormulaeHandler(self.excel_data, self.brand_var.get(), self.artwork_var.get())
        self.excel_data = formulae_handler.apply_filters()
        
        # Determine which of the potential columns exists in the DataFrame
        found_column = None
        for col in potential_columns:
            if col in self.excel_data.columns:
                found_column = col
                break

        if not found_column:
            messagebox.showerror("Error", "None of the potential columns found in Excel data.")
            print(self.excel_data.columns)
            found_column = input('Please enter the potential column')

        # Extract last 5 characters from the unique values of the found column
        last_5_chars = ','.join(self.excel_data[found_column].astype(str).str[-5:].unique())
        
        # Generate default filename based on today's date
        today_date = datetime.now().strftime("%d-%m-%Y")
        default_filename = f"{self.party_var.get()} {self.brand_var.get()} {self.artwork_var.get()} {last_5_chars} {today_date}_A.xlsx"
        default_dir = self.output_dir_label.cget("text")
        initialdir = default_dir if os.path.isdir(default_dir) else None
        save_path = filedialog.asksaveasfilename(initialdir=initialdir, initialfile=default_filename, defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx *.xls")])
        if save_path:
            self.excel_data.to_excel(save_path, index=False)
            self.excel_data.to_excel(f"C:\\Users\\dell\\Desktop\\{default_filename}", index=False)
            self.excel_data.to_excel(r'C:\anurag\exc.xlsx', index=False)
            cdr_file_path = self.change_extension(save_path)
            self.read_excel(cdr_file_path)
            self.create_blank_cdr_file(cdr_file_path)
            
            # Text the file to save
            self.text_excel(save_path)
            self.text_excel(f"C:\\Users\\dell\\Desktop\\{default_filename}")
            
            messagebox.showinfo("Success", "Filtered Excel file saved successfully.")

            # Open directories and files
            try:
                # Open the output directory if it exists
                output_dir = self.output_dir_label.cget("text")
                if os.path.isdir(output_dir):
                    # Convert path to absolute and use raw string to handle special characters
                    output_dir = os.path.abspath(output_dir)
                    subprocess.run(['explorer', output_dir], check=False)
                else:
                    print(f"Output directory does not exist: {output_dir}")

            except Exception as e:
                messagebox.showerror("Error", f"An error occurred while opening files: {e}")
    def text_excel(self, file_path):        
        
        # Load the workbook and select the active sheet
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Create a style for text format
        text_style = NamedStyle(name='text', number_format='@')

        # Identify the column index of "MRP"
        mrp_column_index = None
        for col_idx, cell in enumerate(sheet[1], start=1):  # assuming the header is in the first row
            if cell.value == "MRP":
                mrp_column_index = col_idx
                break
                
        # If "MRP" column is found, iterate through the rows
        if mrp_column_index:
            for row in sheet.iter_rows(min_row=2):  # starting from the second row (assuming header is in first row)
                cell = row[mrp_column_index - 1]  # adjust index to 0-based
                # Set the number format of the cell to text
                cell.style = text_style

                # Append '.00' to MRP values
                if cell.value is not None:
                    cell.value = f"{cell.value}.00"

        # Iterate over all cells in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                # Set the number format of the cell to text
                cell.style = text_style
                # If the cell is not empty and is a number or a date, convert it to a string
                if cell.value is not None:
                    if isinstance(cell.value, (int, float)):
                        # Convert number to string
                        cell.value = str(cell.value)
                    elif isinstance(cell.value, datetime):
                        # Convert date to "mm-yyyy" format
                        cell.value = cell.value.strftime("%m-%Y")

        # Set header row to bold and background color grey
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        for cell in sheet[1]:  # assuming the header is in the first row
            cell.font = header_font
            cell.fill = header_fill

        # Autofit columns
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        # Save the workbook with a new name to avoid overwriting the original file
        wb.save(file_path)

class FormulaeHandler():
    def __init__(self, data, brand, artwork):
        self.data = data
        self.brand = brand
        self.artwork = artwork
    
    def apply_filters(self):
        if self.brand == 'EASY BUY' and self.artwork== 'TAG':
            return self.easy_buy_tag()
        elif self.brand == 'PRAKRITI' and self.artwork=='WASHCARE':
            return self.prakriti_washcare()
        elif self.brand == 'PRAKRITI' and self.artwork=='TAG':
            return self.prakriti_washcare()
        elif self.brand == 'PRAKRITI' and self.artwork=='BARCODE':
            return self.prakriti_washcare()
        elif self.brand == 'SHOPPERS STOP' and self.artwork=='WASHCARE':
            return self.shoppers_stop_washcare()
        elif self.brand == 'EASY BUY' and self.artwork=='TRACIBILITY':
            return self.easy_buy_tracibility()
        elif self.brand == 'KOLKATA BAZAAR' and self.artwork=='WASHCARE':
            return self.kolkata_bazaar()
        elif self.brand == 'PANTALOON' and self.artwork=='WASHCARE':
            return self.pantaloon_washcare()
        elif self.brand == 'U R YOU' and self.artwork=='WASHCARE & BARCODE':
            return self.ur_you_washcare_barcode()
        elif self.brand == 'U R YOU' and self.artwork=='BARCODE':
            return self.ur_you_washcare_barcode()
        elif self.brand == 'HAUTE CURRY' and self.artwork=='WASHCARE':
            return self.shoppers_stop_washcare()
        elif self.brand == 'HAUTE CURRY' and self.artwork=='BARCODE':
            return self.haute_curry_barcode()
        elif self.brand == 'HAUTE CURRY' and self.artwork=='WASHCARE':
            return self.haute_curry_barcode()
        elif self.brand == 'HAUTE CURRY' and self.artwork=='BARCODE AND WASHCARE':
            return self.haute_curry_barcode()
        elif self.brand == 'INTUNE' and self.artwork=='BARCODE':
            return self.intune_barcode()
        elif self.brand == 'INTUNE' and self.artwork=='BARCODE AND WASHCARE':
            return self.intune_barcode()
        else:
            return self.data
            
    def intune_barcode(self):
        # Assuming self.data is a pandas DataFrame or similar structure
        if 'Size' in self.data.columns and 'Size in Cm' in self.data.columns:
            # Convert Size and CM Size columns to strings if they are not already
            self.data['Size'] = self.data['Size'].astype(str)
            self.data['Size in Cm'] = self.data['Size in Cm'].astype(str)
            
            # Create Combined Size column with ' / (' format
            self.data['Size / CM SIZE'] = self.data['Size'] + ' / ' + self.data['Size in Cm']
            
            if 'Vendor Part number' not in self.data.columns:
                self.data['Vendor Part number'] = self.data['Generic/Single article code']
        
        return self.data       
        
    def haute_curry_barcode(self):
        # Assuming self.data is a pandas DataFrame or similar structure
        if 'Size' in self.data.columns and 'CM Size' in self.data.columns:
            # Convert Size and CM Size columns to strings if they are not already
            self.data['Size'] = self.data['Size'].astype(str)
            self.data['CM Size'] = self.data['CM Size'].astype(str)
            
            # Create Combined Size column with ' / (' format
            self.data['Size / (CM SIZE)'] = self.data['Size'] + ' / (' + self.data['CM Size'] + ')'
            
            if 'Vendor Part number' not in self.data.columns:
                self.data['Vendor Part number'] = self.data['Generic/Single article code']
        
        return self.data        
            
    def ur_you_washcare_barcode(self):
        # Assuming self.data is a pandas DataFrame or similar structure
        if 'Size' in self.data.columns and 'CM Size' in self.data.columns:
            # Convert Size and CM Size columns to strings if they are not already
            self.data['Size'] = self.data['Size'].astype(str)
            self.data['CM Size'] = self.data['CM Size'].astype(str)
            
            # Create Combined Size column with ' / (' format
            self.data['Size / (CM SIZE)'] = self.data['Size'] + ' (' + self.data['CM Size'] + ')'
            
            if 'Vendor Part number' not in self.data.columns:
                self.data['Vendor Part number'] = self.data['Generic/Single article code']
        
        return self.data
            
    def pantaloon_washcare(self):
        # Update 'Color Family' column values to uppercase and add ' / '
        self.data['Color Family'] = self.data['Color Family'].str.upper() + ' / '
        # Make 'Color Family' column values uppercase
        self.data['COLOR'] = self.data['COLOR'].str.upper()
        
        return self.data          
    def kolkata_bazaar(self):
        return self.prakriti_washcare()
    def easy_buy_tracibility(self):
        if 'SHORT_DESC:COLOR_DESC:SIZE_DESC' not in self.data.columns:
            # Assuming 'SHORT_DESC', 'COLOR_DESC', 'SIZE_DESC' are existing columns
            if all(col in self.data.columns for col in ['SHORT_DESC', 'COLOR_DESC', 'SIZE_DESC']):
                self.data['SHORT_DESC:COLOR_DESC:SIZE_DESC'] = self.data['SHORT_DESC'] + ':' + \
                                                                self.data['COLOR_DESC'] + ':' + \
                                                                self.data['SIZE_DESC']
            else:
                print("Required columns (SHORT_DESC, COLOR_DESC, SIZE_DESC) are missing.")
        else:
            print("Column 'SHORT_DESC:COLOR_DESC:SIZE_DESC' already exists.")
        
        return self.data
    def split_text(self, column_name, first_column, deliminator, second_column):
        # Assuming self.data is a DataFrame and 'PRODUCT DIMENSION' column exists
        if isinstance(self.data, pd.DataFrame) and 'column_name' in self.data.columns:
            # Split 'PRODUCT DIMENSION' into 'Size' and 'SIZE in CM'
            self.data[[first_column, second_column]] = self.data[column_name].str.split(deliminator, expand=True)
            
            # Strip leading and trailing whitespaces from the new columns
            self.data[first_column] = self.data[first_column].str.strip()
            self.data[second_column] = self.data[second_column].str.strip()
            
            # Return or do something with the modified DataFrame
            return self.data
        else:
            print(f"DataFrame should contain '{column_name}' column.")
            return self.data
    def shoppers_stop_washcare(self):
        # Assuming self.data is a DataFrame and 'PRODUCT DIMENSION' column exists
        if isinstance(self.data, pd.DataFrame) and 'PRODUCT DIMENSION' in self.data.columns:
            # Split 'PRODUCT DIMENSION' into 'Size' and 'SIZE in CM'
            self.data[['Size', 'SIZE in CM']] = self.data['PRODUCT DIMENSION'].str.split('/', expand=True)
            
            # Strip leading and trailing whitespaces from the new columns
            self.data['Size'] = self.data['Size'].str.strip()
            self.data['SIZE in CM'] = self.data['SIZE in CM'].str.strip()
            
            # Return or do something with the modified DataFrame
            return self.data
        else:
            print("DataFrame should contain 'PRODUCT DIMENSION' column.")
            return self.data
    
    def prakriti_washcare(self):
        # Assuming self.data is a DataFrame and 'PRODUCT DIMENSION' column exists
        if isinstance(self.data, pd.DataFrame) and 'PRODUCT DIMENSION' in self.data.columns:
            # Split 'PRODUCT DIMENSION' into 'CODE' and 'SIZE in CM'
            self.data[['CODE', 'SIZE in CM']] = self.data['PRODUCT DIMENSION'].str.split('/', expand=True)
            
            # Strip leading and trailing whitespaces from the new columns
            self.data['CODE'] = self.data['CODE'].str.strip()
            self.data['SIZE in CM'] = self.data['SIZE in CM'].str.strip()
            
            # Return or do something with the modified DataFrame
            return self.data
        else:
            print("DataFrame should contain 'PRODUCT DIMENSION' column.")
            return self.data
    
    def easy_buy_tag(self):
        # Assuming self.data is a DataFrame and 'SEASON' and 'SUPPLIER' columns exist
        if isinstance(self.data, pd.DataFrame) and 'SEASON' in self.data.columns and 'SUPPLIER' in self.data.columns:
            # Convert 'SEASON' and 'SUPPLIER' columns to string type
            self.data['SEASON'] = self.data['SEASON'].astype(str)
            self.data['SUPPLIER'] = self.data['SUPPLIER'].astype(str)
            
            # Combine 'SEASON' and 'SUPPLIER' columns with ' - ' as separator into a new column
            self.data['SEASON - SUPPLIER'] = self.data['SEASON'] + ' - ' + self.data['SUPPLIER']
            
            # Create 'STYLE CODE' column with first 10 characters of 'STYLE_CODE_ITEM_DESCRIPTION'
            self.data['STYLE CODE'] = self.data['STYLE_CODE_ITEM_DESCRIPTION'].str[:10]
            
                
            # Return or do something with the modified DataFrame
            return self.data
        else:
            print("DataFrame should contain 'SEASON' and 'SUPPLIER' columns.")

class MethodsHandler():
    def __init__(self, master):
        self.master = master
        # Retrieve values from the variables
        self.brand = self.master.brand_var.get()
        self.party = self.master.party_var.get()
        self.artwork = self.master.artwork_var.get()
        
    def repeat_command(self):
        # Ask user for the number of times to repeat the process
        try:
            num_iterations = 10
        except ValueError:
            print("Invalid input. Please enter a valid number.")
            return
            
        time.sleep(2)

        # Loop for the specified number of iterations
        for _ in range(num_iterations):
            # Get current mouse position
            current_mouse_x, current_mouse_y = pg.position()

            # Hold down Ctrl key
            pg.keyDown('ctrl')

            # Click at the current mouse position
            pg.click(current_mouse_x, current_mouse_y)

            # Release Ctrl key
            pg.keyUp('ctrl')

            # Press Ctrl+R (reload or refresh depending on the context)
            pg.hotkey('ctrl', 'r')

            # Wait for a brief moment to ensure the page reloads
            time.sleep(0.3)

            # Press Page Down
            pg.press('pagedown')

            # Wait for 0.2 seconds before the next iteration
            time.sleep(0.2)
        
    def print_import(self):
        time.sleep(3)
        pg.hotkey('ctrl', 'p')
        time.sleep(2)
        pg.click(709,572)

        time.sleep(10)
        pg.click(849,649)

        time.sleep(2)
        pg.click(795,441)

        time.sleep(2)
        pg.click(607,745)

        time.sleep(2)
        pg.click(600,400)

        pg.hotkey('ctrl', 'i')

        time.sleep(2)
        pg.click(400,140)
        pg.doubleClick() 

        time.sleep(2)
        pg.press('enter')
        time.sleep(2)
        pg.press('enter')
        time.sleep(2)
        pg.press('enter')
        
    def add_db(self):
        software = self.sample_file_opener()
        if software == "BARTENDER":
            confirm = messagebox.askyesno("File Opened Or Not", "Confirm the sample")
            if confirm:
                self.formulate_bartender()
        elif software =="COREL DRAW":
            confirm = messagebox.askyesno("File Opened Or Not", "Confirm the sample")
            if confirm:
                self.formulate_corel()
        
        
    def sample_file_opener(self):
        json_path = r'\\deepa\d\MAIL-2024\ExcelFilterApp\stps.json'

        try:
            # Open the JSON file and load data
            with open(json_path, 'r') as file:
                data = json.load(file)
                print(data)
        except Exception as e:
            print(f'Error loading the json {e}')
        
        import win32com.client
        file_path = None
        # Connect to CorelDRAW application
        corel_app = win32com.client.Dispatch("CorelDRAW.Application")
        
        file_path = data.get(self.brand, {}).get(self.artwork,{}).get("PATH",None)
        software = data.get(self.brand, {}).get(self.artwork,{}).get("SOFTWARE",None)
        
        if file_path and software == "COREL DRAW":
            corel_doc = corel_app.OpenDocument(file_path)
            return "COREL DRAW"
        elif file_path and software == "BARTENDER":
            os.system(f'start "" "{file_path}"')
            return "BARTENDER"
        else:
            confirm = messagebox.askyesno("File not found", "Do you want to add file")
            if confirm:
                new_data = self.add_file_path(json_path)
                            
    def add_file_path(self, json_path):
        def add_data():
            with open(json_path, 'r') as file:
                data = json.load(file)
            if self.brand in data:
                if self.artwork in data[self.brand]:
                    print(f"The brand {self.brand} and {self.artwork} is available")
                else:
                    data[self.brand][self.artwork]={"PATH": file_path,  "SOFTWARE": software.get()}
            else:
                data[self.brand]={self.artwork:{"PATH": file_path,  "SOFTWARE": software.get()}}
                print(f'Updated values{data}')
                
            with open(json_path, 'w') as file:
                json.dump(data, file, indent = 4)
        # Ask user for file path
        root = ctk.CTk()
        root.geometry("300x100")
        file_path = filedialog.askopenfilename(title="Select a file")
        software = ctk.CTkComboBox(root, values=["BARTENDER", "COREL DRAW"])
        software.pack()
        submit = ctk.CTkButton(root, text= "Submit now", command =lambda:(add_data()))
        submit.pack(pady = '5px')
        root.mainloop()
        
        
    def formulate_bartender(self):
        import pygetwindow as gw
        import pyautogui as pg
        import time

        # Get all currently opened windows
        windows = gw.getAllTitles()
        # Count the number of opened windows
        num_windows = len(windows)
        print(num_windows)
        pg.click(151, 88, duration = 0.5)
        time.sleep(1)

        windows = gw.getAllTitles()
        num_windows1 = len(windows)
        windows = gw.getAllTitles()

        if(num_windows+2==num_windows1):
            print('two opened')
            pg.press('esc')

        time.sleep(1)    
        pg.click(544, 540)
        time.sleep(0.5)
        pg.click(732, 424)
        time.sleep(0.5)
        pg.click(468, 539)
        time.sleep(0.5)
        pg.click(813, 536)
        time.sleep(0.5)
        pg.click(662, 380)
        time.sleep(0.5)
        pg.click(813, 536)
        time.sleep(0.5)
        pg.click(609, 346)
        pg.click(609, 346)
        # DRIVE SELECTION
        time.sleep(1)
        pg.click(850, 509)
        pg.press('down')
        pg.press('down')
        pg.press('down')
        pg.press('down')
        pg.press('enter')
        # C DRIVE
        pg.click(759, 374)
        pg.click(759, 374)
        pg.click(759, 374)
        # Loop to press the down arrow key multiple times
        for _ in range(10):
            pg.press('down')

        pg.press('up')
        pg.press('up')
        pg.press('enter')
        pg.press('down')
        pg.press('enter')
        for _ in range(3):
            pg.press('down')
        pg.press('enter')
        pg.click(668, 372)
        pg.press('enter')

        # select the sheet
        pg.click(902, 321)
        pg.press('enter')

        #click ok
        pg.click(727, 573)
    def formulate_corel(self):
        
        import pygetwindow as gw
        import pyautogui as pg
        import time

        # Get all currently opened windows
        windows = gw.getAllTitles()
        # Count the number of opened windows
        num_windows = len(windows)
        print(num_windows)
        pg.click(766, 58, duration = 0.5)
        # time.sleep(0.5)
        pg.click(662, 380, duration = 0.5)
        pg.click(776, 561)
        pg.click(x=439, y=299, duration = 0.2)
        pg.click(x=659, y=299, duration = 0.2)
        pg.click(x=548, y=189, duration = 0.2)
        pg.click(x=470, y=266, duration = 0.2)
        pg.press('enter')
        # DRIVE SELECTION
        pg.click(x=717, y=472, duration = 0.2)
        pg.press('down')
        pg.press('down')
        pg.press('down')
        pg.press('down')
        pg.press('enter')
        # C DRIVE
        pg.click(x=676, y=336)
        pg.click(x=676, y=336)
        pg.click(x=676, y=336)
        # Loop to press the down arrow key multiple times
        for _ in range(10):
            pg.press('down')
        pg.press('up')
        pg.press('up')
        pg.press('enter')
        pg.press('down')
        pg.press('enter')
        for _ in range(3):
            pg.press('down')
        pg.press('enter')
        pg.click(x=534, y=338)
        pg.press('enter')
        pg.press('enter')
        pg.click(x=752, y=561, duration = 0.2)
        pg.click(x=752, y=561, duration = 0.2)
        pg.click(x=752, y=561, duration = 0.2)
        pg.click(x=752, y=561, duration = 0.2)
        

if __name__ == "__main__":
    app = ExcelFilterApp()
    app.mainloop()
